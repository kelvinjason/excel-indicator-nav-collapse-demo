from __future__ import annotations

from datetime import date
from pathlib import Path
from statistics import mean

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT = Path("indicator_navigation_demo.xlsx")

CATEGORIES = [
    {
        "category": "Traffic",
        "owner": "Growth Team",
        "items": [
            {
                "id": "TRF-001",
                "name": "Website Sessions",
                "definition": "Total qualified sessions from all acquisition channels.",
                "unit": "sessions",
                "current": 128400,
                "target": 135000,
                "trend": "↗ +6.3%",
                "updated_at": "2026-04-08",
                "notes": "Strong uplift from paid social and SEO refresh.",
            },
            {
                "id": "TRF-002",
                "name": "Organic CTR",
                "definition": "Click-through rate from organic search impressions.",
                "unit": "%",
                "current": 4.8,
                "target": 5.2,
                "trend": "↗ +0.4pt",
                "updated_at": "2026-04-08",
                "notes": "Meta title rewrite is improving click quality.",
            },
            {
                "id": "TRF-003",
                "name": "Landing Page Bounce Rate",
                "definition": "Percentage of visitors leaving after viewing one page.",
                "unit": "%",
                "current": 38.2,
                "target": 35.0,
                "trend": "↘ -1.1pt",
                "updated_at": "2026-04-08",
                "notes": "Still above target; hero section content needs cleanup.",
            },
        ],
    },
    {
        "category": "Conversion",
        "owner": "Product Team",
        "items": [
            {
                "id": "CNV-001",
                "name": "Signup Conversion Rate",
                "definition": "Percentage of visitors completing account registration.",
                "unit": "%",
                "current": 7.4,
                "target": 7.0,
                "trend": "↗ +0.6pt",
                "updated_at": "2026-04-08",
                "notes": "New registration flow reduced drop-off on mobile.",
            },
            {
                "id": "CNV-002",
                "name": "Demo Request Rate",
                "definition": "Share of qualified leads requesting a product demo.",
                "unit": "%",
                "current": 2.1,
                "target": 2.5,
                "trend": "→ flat",
                "updated_at": "2026-04-08",
                "notes": "CTA wording is under A/B test this week.",
            },
            {
                "id": "CNV-003",
                "name": "Checkout Completion Rate",
                "definition": "Percentage of carts that complete payment successfully.",
                "unit": "%",
                "current": 81.6,
                "target": 84.0,
                "trend": "↗ +1.2pt",
                "updated_at": "2026-04-08",
                "notes": "Payment retries helped but address validation is still noisy.",
            },
        ],
    },
    {
        "category": "Revenue",
        "owner": "Commercial Ops",
        "items": [
            {
                "id": "REV-001",
                "name": "Monthly Recurring Revenue",
                "definition": "Recurring subscription revenue recognized in the month.",
                "unit": "USD",
                "current": 486000,
                "target": 500000,
                "trend": "↗ +4.9%",
                "updated_at": "2026-04-08",
                "notes": "Enterprise expansion deals are pacing slightly behind plan.",
            },
            {
                "id": "REV-002",
                "name": "Average Order Value",
                "definition": "Average net order value after discounts and returns.",
                "unit": "USD",
                "current": 182.4,
                "target": 175.0,
                "trend": "↗ +3.1%",
                "updated_at": "2026-04-08",
                "notes": "Bundle attach rate improved after pricing refresh.",
            },
            {
                "id": "REV-003",
                "name": "Gross Margin",
                "definition": "Revenue remaining after direct cost of goods sold.",
                "unit": "%",
                "current": 58.7,
                "target": 60.0,
                "trend": "→ flat",
                "updated_at": "2026-04-08",
                "notes": "Fulfillment cost spike dragged margin in week 1.",
            },
        ],
    },
    {
        "category": "Retention",
        "owner": "Customer Success",
        "items": [
            {
                "id": "RET-001",
                "name": "30-Day Retention Rate",
                "definition": "Percentage of new users active again within 30 days.",
                "unit": "%",
                "current": 42.9,
                "target": 45.0,
                "trend": "↗ +1.5pt",
                "updated_at": "2026-04-08",
                "notes": "Lifecycle messaging helped but activation still needs work.",
            },
            {
                "id": "RET-002",
                "name": "Logo Churn Rate",
                "definition": "Percentage of active customers lost during the month.",
                "unit": "%",
                "current": 1.8,
                "target": 1.5,
                "trend": "↘ -0.2pt",
                "updated_at": "2026-04-08",
                "notes": "Improving, but still above the acceptable threshold.",
            },
            {
                "id": "RET-003",
                "name": "Net Promoter Score",
                "definition": "Customer advocacy score from post-interaction surveys.",
                "unit": "score",
                "current": 46,
                "target": 50,
                "trend": "↗ +3",
                "updated_at": "2026-04-08",
                "notes": "Support quality gains are visible in promoter comments.",
            },
        ],
    },
]

THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FILL = PatternFill("solid", fgColor="0F172A")
SECTION_FILL = PatternFill("solid", fgColor="DBEAFE")
HEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
CATEGORY_FILL = PatternFill("solid", fgColor="F8FAFC")
GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")
AMBER_FILL = PatternFill("solid", fgColor="FEF3C7")
RED_FILL = PatternFill("solid", fgColor="FEE2E2")
LINK_FONT = Font(color="0563C1", underline="single")
WHITE_BOLD = Font(color="FFFFFF", bold=True, size=14)
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def compute_status(item: dict) -> str:
    current = item["current"]
    target = item["target"]
    lower_is_better = "Bounce Rate" in item["name"] or "Churn Rate" in item["name"]
    if lower_is_better:
        ratio = target / current if current else 0
    else:
        ratio = current / target if target else 0

    if ratio >= 1:
        return "On Track"
    if ratio >= 0.95:
        return "Watch"
    return "Risk"


for category in CATEGORIES:
    for item in category["items"]:
        item["category"] = category["category"]
        item["owner"] = category["owner"]
        item["status"] = compute_status(item)

all_items = [item for category in CATEGORIES for item in category["items"]]

wb = Workbook()
ws_summary = wb.active
ws_summary.title = "Summary"
ws_detail = wb.create_sheet("Indicators")

# --- Summary sheet ---
ws_summary["A1"] = "Indicator Navigation Demo"
ws_summary["A1"].font = WHITE_BOLD
ws_summary["A1"].fill = TITLE_FILL
ws_summary["A1"].alignment = LEFT
ws_summary.merge_cells("A1:I1")

ws_summary["A2"] = (
    "Click an indicator name to jump to its detail anchor on the Indicators sheet. "
    "Indicator groups are collapsed by default in the detail sheet."
)
ws_summary.merge_cells("A2:I2")
ws_summary["A2"].alignment = LEFT

kpi_headers = ["Category", "Owner", "Indicators", "On Track", "Watch", "Risk", "Avg. Progress"]
kpi_start = 4
for col, header in enumerate(kpi_headers, start=1):
    cell = ws_summary.cell(kpi_start, col, header)
    cell.font = BOLD
    cell.fill = HEADER_FILL
    cell.border = BORDER
    cell.alignment = CENTER

kpi_row = kpi_start + 1
for category in CATEGORIES:
    items = category["items"]
    on_track = sum(1 for x in items if x["status"] == "On Track")
    watch = sum(1 for x in items if x["status"] == "Watch")
    risk = sum(1 for x in items if x["status"] == "Risk")
    progress_values = []
    for x in items:
        lower_is_better = "Bounce Rate" in x["name"] or "Churn Rate" in x["name"]
        if lower_is_better:
            progress_values.append(x["target"] / x["current"] if x["current"] else 0)
        else:
            progress_values.append(x["current"] / x["target"] if x["target"] else 0)
    avg_progress = mean(progress_values)
    values = [
        category["category"],
        category["owner"],
        len(items),
        on_track,
        watch,
        risk,
        avg_progress,
    ]
    for col, value in enumerate(values, start=1):
        cell = ws_summary.cell(kpi_row, col, value)
        cell.border = BORDER
        cell.alignment = CENTER if col != 2 else LEFT
        if col == 7:
            cell.number_format = "0.0%"
    kpi_row += 1

summary_headers = [
    "Category",
    "Indicator ID",
    "Indicator Name (clickable)",
    "Current",
    "Target",
    "Status",
    "Trend",
    "Owner",
    "Last Update",
]
summary_start = 11
for col, header in enumerate(summary_headers, start=1):
    cell = ws_summary.cell(summary_start, col, header)
    cell.font = BOLD
    cell.fill = HEADER_FILL
    cell.border = BORDER
    cell.alignment = CENTER

indicator_anchor_rows: dict[str, int] = {}

# --- Detail sheet ---
ws_detail["A1"] = "Indicator Detail Sheet"
ws_detail["A1"].font = WHITE_BOLD
ws_detail["A1"].fill = TITLE_FILL
ws_detail["A1"].alignment = LEFT
ws_detail.merge_cells("A1:M1")
ws_detail["A2"] = (
    "Each category is collapsed by default. Use Excel's native outline +/- controls on the left to expand/collapse a category block."
)
ws_detail.merge_cells("A2:M2")
ws_detail["A2"].alignment = LEFT
ws_detail.sheet_properties.outlinePr.summaryBelow = False
ws_detail.sheet_view.showOutlineSymbols = True

current_row = 4
for idx, category in enumerate(CATEGORIES, start=1):
    # category summary row
    ws_detail.cell(current_row, 1, "＋")
    ws_detail.cell(current_row, 2, category["category"])
    ws_detail.cell(current_row, 3, f"CAT-{idx:02d}")
    ws_detail.cell(current_row, 4, f"{category['category']} Indicators")
    ws_detail.cell(current_row, 5, "Category summary row. Expand below for the detailed table.")
    ws_detail.cell(current_row, 6, "-")
    ws_detail.cell(current_row, 7, len(category["items"]))
    ws_detail.cell(current_row, 8, "-")
    ws_detail.cell(current_row, 9, "Collapsed by default")
    ws_detail.cell(current_row, 10, "Use outline +/-")
    ws_detail.cell(current_row, 11, str(date.today()))
    ws_detail.cell(current_row, 12, category["owner"])
    ws_detail.cell(current_row, 13, "Tip: click a Summary indicator to land on a specific detail row.")
    for col in range(1, 14):
        cell = ws_detail.cell(current_row, col)
        cell.font = BOLD
        cell.fill = SECTION_FILL
        cell.border = BORDER
        cell.alignment = LEFT if col not in {1, 7, 8, 11} else CENTER
    category_row = current_row
    current_row += 1

    table_header_row = current_row
    detail_headers = [
        "Toggle",
        "Category",
        "Indicator ID",
        "Indicator Name",
        "Definition",
        "Unit",
        "Current",
        "Target",
        "Status",
        "Trend",
        "Update Date",
        "Owner",
        "Notes",
    ]
    for col, header in enumerate(detail_headers, start=1):
        cell = ws_detail.cell(table_header_row, col, header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = CENTER
    current_row += 1

    for item in category["items"]:
        indicator_anchor_rows[item["id"]] = current_row
        values = [
            "－",
            item["category"],
            item["id"],
            item["name"],
            item["definition"],
            item["unit"],
            item["current"],
            item["target"],
            item["status"],
            item["trend"],
            item["updated_at"],
            item["owner"],
            item["notes"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws_detail.cell(current_row, col, value)
            cell.border = BORDER
            cell.alignment = LEFT if col not in {1, 7, 8, 11} else CENTER
            if col in {7, 8}:
                unit = item["unit"]
                if unit == "%":
                    cell.number_format = '0.0'
                elif unit == "USD":
                    cell.number_format = '$#,##0.0'
                elif unit == "sessions":
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.0'
        status_cell = ws_detail.cell(current_row, 9)
        if item["status"] == "On Track":
            status_cell.fill = GREEN_FILL
        elif item["status"] == "Watch":
            status_cell.fill = AMBER_FILL
        else:
            status_cell.fill = RED_FILL
        current_row += 1

    table_end_row = current_row - 1
    table_ref = f"A{table_header_row}:M{table_end_row}"
    table = Table(displayName=f"Tbl_{category['category']}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws_detail.add_table(table)

    ws_detail.row_dimensions.group(table_header_row, table_end_row, hidden=True, outline_level=1)
    ws_detail.row_dimensions[category_row].outlineLevel = 0
    current_row += 1

summary_row = summary_start + 1
for item in all_items:
    values = [
        item["category"],
        item["id"],
        item["name"],
        item["current"],
        item["target"],
        item["status"],
        item["trend"],
        item["owner"],
        item["updated_at"],
    ]
    for col, value in enumerate(values, start=1):
        cell = ws_summary.cell(summary_row, col, value)
        cell.border = BORDER
        cell.alignment = LEFT if col not in {4, 5} else CENTER

    anchor_row = indicator_anchor_rows[item["id"]]
    safe_name = item["name"].replace('"', '""')
    ws_summary.cell(summary_row, 3).value = f'=HYPERLINK("#\'Indicators\'!A{anchor_row}","{safe_name}")'
    ws_summary.cell(summary_row, 3).font = LINK_FONT
    ws_summary.cell(summary_row, 6).alignment = CENTER
    if item["status"] == "On Track":
        ws_summary.cell(summary_row, 6).fill = GREEN_FILL
    elif item["status"] == "Watch":
        ws_summary.cell(summary_row, 6).fill = AMBER_FILL
    else:
        ws_summary.cell(summary_row, 6).fill = RED_FILL

    for col in (4, 5):
        unit = item["unit"]
        cell = ws_summary.cell(summary_row, col)
        if unit == "%":
            cell.number_format = '0.0'
        elif unit == "USD":
            cell.number_format = '$#,##0.0'
        elif unit == "sessions":
            cell.number_format = '#,##0'
        else:
            cell.number_format = '#,##0.0'
    summary_row += 1

summary_table = Table(displayName="SummaryIndicators", ref=f"A{summary_start}:I{summary_row - 1}")
summary_table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
ws_summary.add_table(summary_table)

for sheet in (ws_summary, ws_detail):
    sheet.freeze_panes = "A4"

summary_widths = {
    "A": 14,
    "B": 14,
    "C": 30,
    "D": 12,
    "E": 12,
    "F": 12,
    "G": 12,
    "H": 18,
    "I": 14,
}
detail_widths = {
    "A": 10,
    "B": 14,
    "C": 14,
    "D": 28,
    "E": 42,
    "F": 12,
    "G": 12,
    "H": 12,
    "I": 14,
    "J": 14,
    "K": 14,
    "L": 18,
    "M": 42,
}
for col, width in summary_widths.items():
    ws_summary.column_dimensions[col].width = width
for col, width in detail_widths.items():
    ws_detail.column_dimensions[col].width = width

wb.save(OUTPUT)
print(f"Created {OUTPUT.resolve()}")
